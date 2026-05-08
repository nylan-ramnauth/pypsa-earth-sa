# -*- coding: utf-8 -*-
# SPDX-FileCopyrightText:  PyPSA-Earth and PyPSA-Eur Authors
#
# SPDX-License-Identifier: AGPL-3.0-or-later
"""Common I/O helpers for the ZA Module 04 source audits.

Provides:
- sha256_of_file: stable file content hash
- sha256_of_dir_listing: hash of (relative path, size) tuples for directories
- count_lines_csv: row count for CSV files (header excluded)
- list_xlsx_sheets: sheet names + per-sheet row/column count
- write_geojson: safe geopandas → GeoJSON export
- registry_row: factory for the canonical registry row dict
"""
from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Iterable

import pandas as pd

REGISTRY_COLUMNS = [
    "source_path",
    "tracked_or_external",
    "hash",
    "sheet_or_layer",
    "row_count",
    "port_policy",
    "owning_module",
    "baseline_use",
    "expansion_use",
    "notes",
]


def sha256_of_file(path: Path, chunk: int = 1 << 20) -> str:
    """Return the SHA256 hex digest of a file. Returns empty string on failure."""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as fh:
            for blk in iter(lambda: fh.read(chunk), b""):
                h.update(blk)
        return h.hexdigest()
    except Exception:
        return ""


def sha256_of_dir_listing(path: Path) -> str:
    """Return a deterministic hash of (relative path, size) pairs under *path*."""
    base = Path(path)
    if not base.exists() or not base.is_dir():
        return ""
    entries: list[tuple[str, int]] = []
    for p in sorted(base.rglob("*")):
        if p.is_file():
            try:
                entries.append((str(p.relative_to(base)), p.stat().st_size))
            except Exception:
                continue
    h = hashlib.sha256()
    for rel, size in entries:
        h.update(f"{rel}|{size}\n".encode("utf-8"))
    return h.hexdigest()


def count_lines_csv(path: Path) -> int:
    """Row count for a CSV file, excluding the header. Returns -1 on failure."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as fh:
            r = csv.reader(fh)
            n = sum(1 for _ in r)
        return max(n - 1, 0)
    except Exception:
        return -1


def list_xlsx_sheets(path: Path) -> list[dict]:
    """Return [{sheet, n_rows, n_cols, columns}] for an xlsx workbook."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        out = []
        for name in wb.sheetnames:
            ws = wb[name]
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0
            cols: list[str] = []
            if n_rows >= 1 and n_cols >= 1:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    cols = [str(c) if c is not None else "" for c in first_row]
            out.append({"sheet": name, "n_rows": int(n_rows), "n_cols": int(n_cols), "columns": cols})
        wb.close()
        return out
    except Exception as exc:
        return [{"sheet": "<error>", "n_rows": -1, "n_cols": -1, "columns": [str(exc)]}]


def write_geojson(gdf, out_path: Path) -> int:
    """Write a GeoDataFrame to GeoJSON, returning the feature count. 0 if empty."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if gdf is None or len(gdf) == 0:
        # Still emit an empty FeatureCollection so downstream gates pass.
        out_path.write_text(
            '{"type": "FeatureCollection", "features": []}\n', encoding="utf-8"
        )
        return 0
    try:
        if gdf.crs is None:
            gdf = gdf.set_crs(4326, allow_override=True)
        else:
            gdf = gdf.to_crs(4326)
    except Exception:
        pass
    gdf.to_file(out_path, driver="GeoJSON")
    return int(len(gdf))


def registry_row(
    source_path: str,
    tracked_or_external: str,
    file_hash: str,
    sheet_or_layer: str,
    row_count: int,
    port_policy: str,
    owning_module: str,
    baseline_use: str,
    expansion_use: str,
    notes: str = "",
) -> dict:
    """Build a single canonical registry row."""
    return {
        "source_path": source_path,
        "tracked_or_external": tracked_or_external,
        "hash": file_hash,
        "sheet_or_layer": sheet_or_layer,
        "row_count": row_count,
        "port_policy": port_policy,
        "owning_module": owning_module,
        "baseline_use": baseline_use,
        "expansion_use": expansion_use,
        "notes": notes,
    }


def write_registry_csv(rows: Iterable[dict], out_path: Path) -> int:
    """Write a list of registry rows to CSV with the canonical column order."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(list(rows), columns=REGISTRY_COLUMNS)
    df.to_csv(out_path, index=False)
    return len(df)
